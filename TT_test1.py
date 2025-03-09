import tkinter as tk
from tkinter import messagebox, ttk
import pandas as pd
import random
from tkinter import filedialog
from datetime import datetime
import os
from openpyxl.styles import PatternFill

# Global variables
all_semester_data = []  # Store data for all semesters
time_slots = [
    '9:00-9:55', '9:55-10:50', '11:05-12:00', '12:00-12:55', '13:45-14:40', '14:40-15:35', '15:35-16:30'
]
theory_slots = [
    '9:00-9:55', '10:00-10:50', '11:05-12:00', '12:00-12:55',
    '13:45-14:40', '14:40-15:35', '15:35-16:30'
]
practical_slots = [
    ('9:00-10:50', 2), ('11:05-12:55', 2), ('13:45-15:35', 2), ('14:40-16:30', 2)
]
breaks = [('10:50-11:05', 'Short Break'), ('12:55-13:45', 'Lunch Break')]
end_time_options = ['10:50', '12:55', '14:40', '15:35', '16:30']

# Define colors for different types of classes
COLORS = {
    'theory': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # Green
    'practical': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # Red
    'tutorial': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # Blue
}

class SemesterGUI:
    def __init__(self):
        # Initialize the main window
        self.window = tk.Tk()
        self.window.title("Semester Timetable Generator")
        self.window.geometry("1000x700")
        self.window.minsize(800, 600)
        self.window.configure(bg="#f7f7f7")

        # Initialize the data structures
        self.subjects_entries = []
        self.teachers_entries = []
        self.credits_entries = []
        self.semester_data = {
            'subjects': [],
            'teachers': [],
            'credits': [],
            'semester': '',
            'term_start': '',
            'term_end': '',
            'room_number': '',
            'num_students': 0,
            'file_location': '',
            'excel_name': ''
        }

        # Setup the GUI
        self.setup_gui()

    def setup_gui(self):
        # Create main scrollable canvas
        self.canvas = tk.Canvas(self.window, bg="#f7f7f7")
        self.scrollbar = ttk.Scrollbar(self.window, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg="#f7f7f7")

        # Configure the scrollable frame
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        # Create the window in the canvas
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw", width=self.window.winfo_screenwidth())
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Pack scrollbar and canvas
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        # Bind mouse wheel scrolling
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        # Top frame for basic inputs
        self.top_frame = tk.Frame(self.scrollable_frame, bg="#f7f7f7")
        self.top_frame.pack(fill=tk.X, padx=20, pady=20)

        # Add basic input fields
        self._setup_basic_inputs()

        # Main frame for subject entries
        self.main_frame = tk.Frame(self.scrollable_frame, bg="#f7f7f7")
        self.main_frame.pack(fill=tk.X, padx=20, pady=10)

        # Add subject button
        self.add_subject_button = tk.Button(
            self.scrollable_frame,
            text="Add Subject",
            command=self.add_subject_fields,
            bg="#4CAF50",
            fg="white",
            padx=10
        )
        self.add_subject_button.pack(pady=5)

        # End time frame
        self.end_time_frame = tk.Frame(self.scrollable_frame, bg="#f7f7f7")
        self.end_time_frame.pack(fill=tk.X, padx=20, pady=10)

        # Setup end time selection
        self._setup_end_time_selection()

        # File location frame
        self._setup_file_location()

        # Button frame at bottom
        self._setup_buttons()

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _setup_basic_inputs(self):
        # Semester, term start, term end, room number, and number of students
        self.top_frame.grid_columnconfigure(1, weight=1)
        self.top_frame.grid_columnconfigure(3, weight=1)
        self.top_frame.grid_columnconfigure(5, weight=1)

        semester_label = tk.Label(self.top_frame, text="Semester:", bg="#f7f7f7")
        semester_label.grid(row=0, column=0, padx=5, sticky='e')
        self.semester_entry = tk.Entry(self.top_frame)
        self.semester_entry.grid(row=0, column=1, padx=5, sticky='ew')

        term_start_label = tk.Label(self.top_frame, text="Term Start (dd/mm/yyyy):", bg="#f7f7f7")
        term_start_label.grid(row=0, column=2, padx=5, sticky='e')
        self.term_start_entry = tk.Entry(self.top_frame)
        self.term_start_entry.grid(row=0, column=3, padx=5, sticky='ew')

        term_end_label = tk.Label(self.top_frame, text="Term End (dd/mm/yyyy):", bg="#f7f7f7")
        term_end_label.grid(row=0, column=4, padx=5, sticky='e')
        self.term_end_entry = tk.Entry(self.top_frame)
        self.term_end_entry.grid(row=0, column=5, padx=5, sticky='ew')

        room_number_label = tk.Label(self.top_frame, text="Room Number:", bg="#f7f7f7")
        room_number_label.grid(row=1, column=0, padx=5, pady=10, sticky='e')
        self.room_number_entry = tk.Entry(self.top_frame)
        self.room_number_entry.grid(row=1, column=1, padx=5, sticky='ew')

        num_students_label = tk.Label(self.top_frame, text="Number of Students:", bg="#f7f7f7")
        num_students_label.grid(row=1, column=2, padx=5, sticky='e')
        self.num_students_entry = tk.Entry(self.top_frame)
        self.num_students_entry.grid(row=1, column=3, padx=5, sticky='ew')

    def _setup_end_time_selection(self):
        self.day_end_time_vars = {}
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

        # Configure grid columns for even spacing
        self.end_time_frame.grid_columnconfigure(1, weight=1)

        for i, day in enumerate(days):
            label = tk.Label(self.end_time_frame, text=f"End Time for {day}:", bg="#f7f7f7")
            label.grid(row=i, column=0, padx=5, pady=5, sticky='e')

            self.day_end_time_vars[day] = tk.StringVar(value='10:50')
            end_time_menu = ttk.Combobox(
                self.end_time_frame,
                textvariable=self.day_end_time_vars[day],
                values=end_time_options,
                width=15
            )
            end_time_menu.grid(row=i, column=1, padx=5, pady=5, sticky='w')

    def _setup_file_location(self):
        file_frame = tk.Frame(self.scrollable_frame, bg="#f7f7f7")
        file_frame.pack(fill=tk.X, padx=20, pady=10)

        file_location_label = tk.Label(file_frame, text="File Location:", bg="#f7f7f7")
        file_location_label.pack()
        self.file_location_entry = tk.Entry(file_frame, width=50)
        self.file_location_entry.pack(pady=5)

        self.browse_button = tk.Button(
            file_frame,
            text="Browse",
            command=self.browse_location,
            bg="#2196F3",
            fg="white",
            padx=5
        )
        self.browse_button.pack(pady=5)

        excel_name_label = tk.Label(file_frame, text="Name of Excel Sheet:", bg="#f7f7f7")
        excel_name_label.pack()
        self.excel_name_entry = tk.Entry(file_frame, width=50)
        self.excel_name_entry.pack(pady=5)

    def _setup_buttons(self):
        button_frame = tk.Frame(self.scrollable_frame, bg="#f7f7f7")
        button_frame.pack(pady=20)

        self.generate_button = tk.Button(
            button_frame,
            text="Generate Timetable",
            command=self.generate_timetables,
            bg="#2196F3",
            fg="white",
            padx=10
        )
        self.generate_button.pack(side=tk.LEFT, padx=10)

        self.new_sem_button = tk.Button(
            button_frame,
            text="New Semester",
            command=self.create_new_semester,
            bg="#FF9800",
            fg="white",
            padx=10
        )
        self.new_sem_button.pack(side=tk.LEFT, padx=10)

        self.end_button = tk.Button(
            button_frame,
            text="End Session",
            command=self.window.quit,
            bg="#F44336",
            fg="white",
            padx=10
        )
        self.end_button.pack(side=tk.LEFT, padx=10)

    def browse_location(self):
        folder_selected = filedialog.askdirectory()
        self.file_location_entry.delete(0, tk.END)
        self.file_location_entry.insert(0, folder_selected)

    def add_subject_fields(self):
        subject_frame = tk.Frame(self.main_frame)
        subject_frame.pack(pady=5)

        subject_label = tk.Label(subject_frame, text="Subject Name:")
        subject_label.grid(row=0, column=0)

        teacher_label = tk.Label(subject_frame, text="Teacher Name:")
        teacher_label.grid(row=0, column=1)

        credits_label = tk.Label(subject_frame, text="Credits (theory:tutorial:practical):")
        credits_label.grid(row=0, column=2)

        subject_entry = tk.Entry(subject_frame)
        subject_entry.grid(row=1, column=0)
        self.subjects_entries.append(subject_entry)

        teacher_entry = tk.Entry(subject_frame)
        teacher_entry.grid(row=1, column=1)
        self.teachers_entries.append(teacher_entry)

        credits_entry = tk.Entry(subject_frame)
        credits_entry.grid(row=1, column=2)
        self.credits_entries.append(credits_entry)

    def save_semester_data(self):
        # Clear previous data
        self.semester_data = {
            'subjects': [],
            'teachers': [],
            'credits': [],
            'semester': '',
            'term_start': '',
            'term_end': '',
            'room_number': '',
            'num_students': 0,
            'file_location': '',
            'excel_name': ''
        }

        # Collect all current semester data
        self.semester_data['semester'] = self.semester_entry.get()
        self.semester_data['term_start'] = self.term_start_entry.get()
        self.semester_data['term_end'] = self.term_end_entry.get()
        self.semester_data['room_number'] = self.room_number_entry.get()
        self.semester_data['num_students'] = self.num_students_entry.get()
        self.semester_data['file_location'] = self.file_location_entry.get()
        self.semester_data['excel_name'] = self.excel_name_entry.get()

        # Collect subjects data
        for i in range(len(self.subjects_entries)):
            self.semester_data['subjects'].append(self.subjects_entries[i].get())
            self.semester_data['teachers'].append(self.teachers_entries[i].get())
            self.semester_data['credits'].append(self.credits_entries[i].get())

        # Validate file location and name
        if not self.semester_data['file_location'] or not self.semester_data['excel_name']:
            messagebox.showerror("Error", "Please provide both file location and excel name")
            return False

        # Add to global data
        all_semester_data.append(self.semester_data.copy())
        return True

    def create_new_semester(self):
        if self.save_semester_data():
            self.window.destroy()  # Close current window
            SemesterGUI()  # Create new semester window

    def generate_timetables(self):  # Retaining original function name
        if self.save_semester_data():
            teacher_schedules = {}
            try:
                for semester in all_semester_data:
                    timetable = self.create_timetable(semester, teacher_schedules)
                    self.save_timetable_to_excel(timetable, semester)
                messagebox.showinfo("Success", "All timetables generated successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Error generating timetables: {str(e)}")

    def create_timetable(self, semester_data, teacher_schedules):
        timetable = {day: {slot: '' for slot in time_slots} 
                for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']}
        # Identify pure practical subjects (0:0:X credits)
        pure_practical_subjects = []
        for i, subject in enumerate(semester_data['subjects']):
            credits = semester_data['credits'][i].split(':')
            try:
                credit_values = [int(c.strip()) for c in credits]
            except ValueError:
                raise ValueError(f"Invalid credit format for subject {subject}: {semester_data['credits'][i]}")

            if len(credit_values) == 3 and credit_values[0] == 0 and credit_values[1] == 0 and credit_values[2] > 0:
                pure_practical_subjects.append((subject, semester_data['teachers'][i]))

        # Schedule pure practical subjects simultaneously
        if pure_practical_subjects:
            self.schedule_simultaneous_practicals(pure_practical_subjects, timetable, teacher_schedules)

        # Schedule practical classes for other subjects
        for i, subject in enumerate(semester_data['subjects']):
            credits = semester_data['credits'][i].split(':')
            try:
                credit_values = [int(c.strip()) for c in credits]
            except ValueError:
                raise ValueError(f"Invalid credit format for subject {subject}: {semester_data['credits'][i]}")

            if len(credit_values) == 3 and credit_values[2] > 0 and (subject, semester_data['teachers'][i]) not in pure_practical_subjects:
                self.schedule_practical(subject, semester_data['teachers'][i], timetable, teacher_schedules)

        # Schedule theory and tutorial classes
        for i, subject in enumerate(semester_data['subjects']):
            credits = semester_data['credits'][i].split(':')
            try:
                credit_values = [int(c.strip()) for c in credits]
            except ValueError:
                raise ValueError(f"Invalid credit format for subject {subject}: {semester_data['credits'][i]}")

            # Schedule theory classes
            if credit_values[0] > 0:
                self.schedule_theory(subject, semester_data['teachers'][i], credit_values[0], timetable, teacher_schedules)
            
            # Schedule tutorial classes
            if len(credit_values) > 1 and credit_values[1] > 0:
                self.schedule_tutorial(subject, semester_data['teachers'][i], credit_values[1], timetable, teacher_schedules)
            self.fill_remaining_slots(timetable, semester_data, teacher_schedules)

            return timetable
    

    # Updated Functions for Timetable Generation


    def schedule_practical(self, subject, teacher, timetable, teacher_schedules):  # Retaining original function name
        days = list(timetable.keys())
        random.shuffle(days)
        credits = [int(c) for c in subject.split(':')] if ':' in subject else [0, 0, 1]
        theory, tutorial, practical = credits

        if theory > 0 and practical > 0:  # For n:0:m credits
            for day in days:
                for slot, duration in practical_slots:
                    if self.is_slot_available(day, slot, teacher, timetable, teacher_schedules):
                        timetable[day][slot] = f"{subject} (Lab - Batch 1) - {teacher}"
                        self.update_teacher_schedule(teacher, day, slot, teacher_schedules)
                        # Ensure no other lab is scheduled simultaneously
                        for other_subject in timetable[day]:
                            if timetable[day][other_subject] and slot != other_subject:
                                if 'Lab' in timetable[day][other_subject]:
                                    return False
                        return True
        else:  # For 0:0:1 format
            self.schedule_simultaneous_practicals([(subject, teacher)], timetable, teacher_schedules)
        return False
    
    def update_teacher_schedule(self, teacher, day, slot, teacher_schedules):  # Retaining original function name
        if teacher not in teacher_schedules:
            teacher_schedules[teacher] = {}
        if day not in teacher_schedules[teacher]:
            teacher_schedules[teacher][day] = []
        teacher_schedules[teacher][day].append(slot)
        
    def fill_remaining_slots(self, timetable, semester_data, teacher_schedules):
        days = list(timetable.keys())
        subjects = semester_data['subjects']
        teachers = semester_data['teachers']

        for day in days:
            end_time = datetime.strptime(self.day_end_time_vars[day].get(), '%H:%M')
            available_slots = [slot for slot in time_slots 
                            if datetime.strptime(slot.split('-')[1], '%H:%M') <= end_time]

            for slot in available_slots:
                if not timetable[day][slot]:
                    # Find a random subject and teacher to fill the slot
                    random_index = random.randint(0, len(subjects) - 1)
                    subject = subjects[random_index]
                    teacher = teachers[random_index]

                    # Schedule a theory class in the empty slot
                    if self.is_slot_available(day, slot, teacher, timetable, teacher_schedules):
                        timetable[day][slot] = f"{subject} (Theory) - {teacher}"
                        
                        # Update teacher schedule
                        if teacher not in teacher_schedules:
                            teacher_schedules[teacher] = {}
                        if day not in teacher_schedules[teacher]:
                            teacher_schedules[teacher][day] = []
                        teacher_schedules[teacher][day].append(slot)

    def schedule_theory(self, subject, teacher, num_classes, timetable, teacher_schedules):
        days = list(timetable.keys())
        random.shuffle(days)
        classes_scheduled = 0

        while classes_scheduled < num_classes:
            for day in days:
                if classes_scheduled >= num_classes:
                    break

                # Get available slots based on end time
                end_time = datetime.strptime(self.day_end_time_vars[day].get(), '%H:%M')
                available_slots = [slot for slot in time_slots 
                                if datetime.strptime(slot.split('-')[1], '%H:%M') <= end_time]
                random.shuffle(available_slots)

                for slot in available_slots:
                    if classes_scheduled >= num_classes:
                        break

                    # Pass teacher_schedules to is_slot_available
                    if self.is_slot_available(day, slot, teacher, timetable, teacher_schedules):
                        timetable[day][slot] = f"{subject} (Theory) - {teacher}"
                        
                        # Update teacher schedule
                        if teacher not in teacher_schedules:
                            teacher_schedules[teacher] = {}
                        if day not in teacher_schedules[teacher]:
                            teacher_schedules[teacher][day] = []
                        teacher_schedules[teacher][day].append(slot)
                        classes_scheduled += 1
                        break

    # Modified schedule_tutorial function and related adjustments
    def schedule_tutorial(self, subject, teacher, num_classes, timetable, teacher_schedules):
        """Schedules tutorial classes as 2-hour blocks (consecutive slots)"""
        days = list(timetable.keys())
        random.shuffle(days)
        classes_scheduled = 0

        while classes_scheduled < num_classes:
            for day in days:
                if classes_scheduled >= num_classes:
                    break

            # Find consecutive slots
                for i in range(len(time_slots)-1):
                    slot1 = time_slots[i]
                    slot2 = time_slots[i+1]
                
                    # Check both slots availability
                    if (self.is_slot_available(day, slot1, teacher, timetable, teacher_schedules) and 
                        self.is_slot_available(day, slot2, teacher, timetable, teacher_schedules)):
                    
                    # Check adjacent classes for both slots
                        if not self.has_adjacent_classes(day, slot1.split('-')[0], teacher, teacher_schedules):
                        # Schedule as 2-hour block
                            timetable[day][slot1] = f"{subject} (Tutorial) - {teacher}"
                            timetable[day][slot2] = f"{subject} (Tutorial) - {teacher}"
                        
                        # Update teacher schedule
                        if teacher not in teacher_schedules:
                            teacher_schedules[teacher] = {}
                        if day not in teacher_schedules[teacher]:
                            teacher_schedules[teacher][day] = []
                        teacher_schedules[teacher][day].extend([slot1, slot2])
                        classes_scheduled += 1
                        break

    def create_timetable(self, semester_data, teacher_schedules):
        timetable = {day: {slot: '' for slot in time_slots} 
                for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']}
    
        # Schedule pure practicals first
        pure_practical_subjects = []
        for i, subject in enumerate(semester_data['subjects']):
            credits = semester_data['credits'][i].split(':')
            credit_values = [int(c.strip()) for c in credits]
            if len(credit_values) == 3 and credit_values[0] == 0 and credit_values[1] == 0 and credit_values[2] > 0:
                pure_practical_subjects.append((subject, semester_data['teachers'][i]))

        if pure_practical_subjects:
            self.schedule_simultaneous_practicals(pure_practical_subjects, timetable, teacher_schedules)

        # Schedule tutorials and practicals with same priority
        for i, subject in enumerate(semester_data['subjects']):
            credits = semester_data['credits'][i].split(':')
            credit_values = [int(c.strip()) for c in credits]
        
        # Schedule tutorials first
            if len(credit_values) > 1 and credit_values[1] > 0:
                self.schedule_tutorial(subject, semester_data['teachers'][i], credit_values[1], timetable, teacher_schedules)
        
        # Schedule non-pure practicals
        if len(credit_values) == 3 and credit_values[2] > 0 and (subject, semester_data['teachers'][i]) not in pure_practical_subjects:
            self.schedule_practical(subject, semester_data['teachers'][i], timetable, teacher_schedules)

        # Schedule theory classes last
        for i, subject in enumerate(semester_data['subjects']):
            credits = semester_data['credits'][i].split(':')
            credit_values = [int(c.strip()) for c in credits]
            if credit_values[0] > 0:
                self.schedule_theory(subject, semester_data['teachers'][i], credit_values[0], timetable, teacher_schedules)

        self.fill_remaining_slots(timetable, semester_data, teacher_schedules)
        return timetable

    def schedule_simultaneous_practicals(self, practical_subjects, timetable, teacher_schedules):
        days = list(timetable.keys())
        random.shuffle(days)
        sessions_needed = 3  # Schedule 3 sessions per week for each batch
        sessions_scheduled = 0

        while sessions_scheduled < sessions_needed:
            for day in days:
                if sessions_scheduled >= sessions_needed:
                    break

                # Try to find two consecutive slots
                for i in range(len(time_slots) - 1):
                    slot1 = time_slots[i]
                    slot2 = time_slots[i + 1]
                    
                    # Pass teacher_schedules to is_slot_available
                    all_teachers_available = all(
                        self.is_slot_available(day, slot1, teacher, timetable, teacher_schedules) and
                        self.is_slot_available(day, slot2, teacher, timetable, teacher_schedules)
                        for _, teacher in practical_subjects
                    )

                    if all_teachers_available:
                        # Check if all teachers have no adjacent classes
                        slot1_start = slot1.split('-')[0]
                        all_no_adjacent_classes = all(
                            not self.has_adjacent_classes(day, slot1_start, teacher, teacher_schedules)
                            for _, teacher in practical_subjects
                        )

                        if all_no_adjacent_classes:
                            # Schedule all practical subjects in these slots
                            batch_num = sessions_scheduled + 1
                            combined_class = []
                            for subject, teacher in practical_subjects:
                                combined_class.append(f"{subject} (Lab - Batch {batch_num}) - {teacher}")
                            
                            combined_label = "\n".join(combined_class)
                            timetable[day][slot1] = combined_label
                            timetable[day][slot2] = combined_label

                            # Update teacher schedules
                            for _, teacher in practical_subjects:
                                if teacher not in teacher_schedules:
                                    teacher_schedules[teacher] = {}
                                if day not in teacher_schedules[teacher]:
                                    teacher_schedules[teacher][day] = []
                                teacher_schedules[teacher][day].extend([slot1, slot2])

                            sessions_scheduled += 1
                            break

    def is_slot_available(self, day, slot, teacher, timetable, teacher_schedules):  # Retaining original function name
        if timetable[day][slot]:
            return False
        if teacher in teacher_schedules and day in teacher_schedules[teacher]:
            for existing_slot in teacher_schedules[teacher][day]:
                if self.is_time_conflict(slot, existing_slot):
                    return False
        return True

    def has_adjacent_classes(self, day, slot_time, teacher, teacher_schedules):
        """
        Checks if the teacher has any classes immediately before or after the given slot.
        
        Args:
            day: The day of the week
            slot_time: Start time of the slot in HH:MM format
            teacher: The teacher to check
            teacher_schedules: Dictionary of all teacher schedules
            
        Returns:
            bool: True if there are adjacent classes, False otherwise
        """
        if teacher not in teacher_schedules or day not in teacher_schedules[teacher]:
            return False
                
        # Find the matching time slot to get the end time
        slot_end = None
        for time_slot in time_slots:
            if time_slot.startswith(slot_time):
                slot_end = datetime.strptime(time_slot.split('-')[1], '%H:%M')
                break
        
        if slot_end is None:
            return False  # Invalid slot time
                
        slot_start = datetime.strptime(slot_time, '%H:%M')
        
        for existing_slot in teacher_schedules[teacher][day]:
            existing_start = datetime.strptime(existing_slot.split('-')[0], '%H:%M')
            existing_end = datetime.strptime(existing_slot.split('-')[1], '%H:%M')
                
            # Check if there's less than a 5-minute break between classes
            if abs((slot_start - existing_end).total_seconds()) < 300 or \
            abs((existing_start - slot_end).total_seconds()) < 300:
                return True
                    
        return False
    def check_teacher_overlap(self, teacher, slot, teacher_schedules):
        for semester_schedule in teacher_schedules.get(teacher, {}).values():
            for scheduled_slot in semester_schedule:
                if self.is_time_conflict(slot, scheduled_slot):
                    return True
        return False
    
    def is_time_conflict(self, slot1, slot2):
        start1, end1 = map(lambda x: datetime.strptime(x, '%H:%M'), slot1.split('-'))
        start2, end2 = map(lambda x: datetime.strptime(x, '%H:%M'), slot2.split('-'))
        return not (end1 <= start2 or start1 >= end2)

    def save_timetable_to_excel(self, timetable, semester_data):
        try:
            # Convert timetable dictionary to DataFrame with time slots as columns and days as rows
            df = pd.DataFrame.from_dict(timetable, orient='index', columns=time_slots)
            
            # Create the full filepath
            filepath = os.path.join(semester_data['file_location'], f"{semester_data['excel_name']}.xlsx")
            
            # Add metadata at the top
            metadata = pd.DataFrame({
                'Semester': [semester_data['semester']],
                'Room Number': [semester_data['room_number']],
                'Number of Students': [semester_data['num_students']],
                'Term Start': [semester_data['term_start']],
                'Term End': [semester_data['term_end']]
            })
            
            # Write to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                metadata.to_excel(writer, sheet_name='Timetable', index=False)
                df.to_excel(writer, sheet_name='Timetable', startrow=len(metadata) + 2)
                
                # Get the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets['Timetable']
                
                # Format the cells
                for col in range(len(df.columns) + 1):  # +1 for index
                    worksheet.column_dimensions[chr(65 + col)].width = 20
                    
                for row in range(len(df) + len(metadata) + 2):
                    worksheet.row_dimensions[row + 1].height = 30
                    
                # Apply colors based on class type
                for row in worksheet.iter_rows(min_row=len(metadata) + 3, max_row=len(df) + len(metadata) + 2, min_col=2, max_col=len(df.columns) + 1):
                    for cell in row:
                        if cell.value:
                            if '(Theory)' in cell.value:
                                cell.fill = COLORS['theory']
                            elif '(Lab)' in cell.value or '(Lab - Batch' in cell.value:
                                cell.fill = COLORS['practical']
                            elif '(Tutorial)' in cell.value:
                                cell.fill = COLORS['tutorial']
                            
            # Ensure properties from previous implementation are retained
            messagebox.showinfo("Success", f"Timetable saved to {filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save timetable: {str(e)}")


def main():
    app = SemesterGUI()
    app.window.mainloop()

if __name__ == "__main__":
    main()